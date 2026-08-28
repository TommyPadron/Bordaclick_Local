# ==============================================================================
# BORDACLICK - APLICACIÓN MÓVIL (ENTORNO DEV)
# Archivo Principal: app_mobile.py
# Descripción: Interfaz web en Streamlit para toma de pedidos, administración y reportes.
# ==============================================================================

# ------------------------------------------------------------------------------
# 1. IMPORTACIÓN DE LIBRERÍAS Y MÓDULOS DEL PROYECTO
# ------------------------------------------------------------------------------
import streamlit as st
import pandas as pd
from datetime import date, timedelta
import plotly.express as px
import plotly.graph_objects as go

# Funciones de la base de datos (SQLite)
try:
    from db_handler import (
        crear_bd,
        obtener_colegios,
        guardar_colegio,
        eliminar_colegio,
        obtener_tipos_prenda,
        guardar_tipo_prenda,
        eliminar_tipo_prenda,
        obtener_tallas,
        guardar_talla,
        eliminar_talla,
        obtener_marcas,
        guardar_marca,
        eliminar_marca,
        obtener_colores,
        guardar_color,
        eliminar_color,
        obtener_zonas_delivery,
        guardar_zona_delivery,
        eliminar_zona_delivery,
        obtener_costo_delivery,
        obtener_parametro,
        guardar_parametro,
        obtener_precio_colegio,
        guardar_orden,
        guardar_detalle,
        obtener_orden_por_id,
        obtener_detalle_orden,
        enviar_pdf_por_correo,
        enviar_confirmacion_solicitud,
        obtener_ordenes,
        registrar_pago,
        obtener_historico_pagos,
        actualizar_status_orden,
        enviar_notificacion_estado,
        eliminar_orden
    )
except ImportError as e:
    st.error(f"Error crítico al importar 'db_handler': {e}")

# Importación segura de herramientas de PDF y Excel con control de excepciones
try:
    from pdf_tools import (
        generar_pdf_orden,
        generar_excel_orden,
        generar_excel_historico
    )
    HERRAMIENTAS_REPORTES_DISPONIBLES = True
except ImportError:
    HERRAMIENTAS_REPORTES_DISPONIBLES = False

# ------------------------------------------------------------------------------
# 2. INICIALIZACIÓN DE BASE DE DATOS Y CONFIGURACIÓN DE PÁGINA
# ------------------------------------------------------------------------------
try:
    crear_bd()
except Exception:
    pass

st.set_page_config(
    page_title="Bordaclick | Bordados y Personalización",
    page_icon="🧵",
    layout="centered"
)

# ------------------------------------------------------------------------------
# 3. CONTROL DE ACCESO (ADMINISTRADOR) Y BARRA LATERAL
# ------------------------------------------------------------------------------
clave_admin = st.sidebar.text_input("Clave Administrador", type="password")

opciones_menu = ["🌐 Inicio / Web", "📝 Nueva Solicitud"]

try:
    clave_correcta = st.secrets["admin"]["password"]
except Exception:
    clave_correcta = "BordaAdmin2026*"

if clave_admin == clave_correcta:
    opciones_menu.extend([
        "📋 Consultas",
        "📊 Reportes",
        "⚙️ Configuración",
        "🏫 Colegios",
        "🚚 Delivery",
        "📦 Prendas",
        "🏷️ Marcas",
        "📏 Tallas",
        "🎨 Colores",
        "💾 Respaldo"
    ])

if "pagina_activa" not in st.session_state:
    st.session_state.pagina_activa = opciones_menu[0]

pagina = st.sidebar.selectbox(
    "Menú", 
    opciones_menu, 
    index=opciones_menu.index(st.session_state.pagina_activa) if st.session_state.pagina_activa in opciones_menu else 0
)
st.session_state.pagina_activa = pagina

# ------------------------------------------------------------------------------
# 4. MEMORIA DE SESIÓN (SESSION STATE)
# ------------------------------------------------------------------------------
if "paso" not in st.session_state:
    st.session_state.paso = 1

if "solicitud_enviada" not in st.session_state:
    st.session_state.solicitud_enviada = False

if "colegios_agregados" not in st.session_state:
    st.session_state.colegios_agregados = []

if "form_version" not in st.session_state:
    st.session_state.form_version = 0

# ------------------------------------------------------------------------------
# 5. ENCABEZADO PRINCIPAL DE LA APLICACIÓN
# ------------------------------------------------------------------------------
col1, col2 = st.columns([1, 3])
with col1:
    try:
        st.image("Logo Bordaclick.JPG", width=80)
    except Exception:
        st.write("🧵")

with col2:
    st.title("🧵 Bordaclick")
    st.caption("Bordados y Personalización Textil Profesional")

st.divider()

# ==============================================================================
# MÓDULO 0: PÁGINA DE INICIO / PRESENTACIÓN WEB
# ==============================================================================
if pagina == "🌐 Inicio / Web":
    
    st.markdown(
        """
        <div style="padding: 10px 0; text-align: center;">
            <h2 style="color: #1E3A8A; margin-bottom: 2px;">🧵 Hilos de Alegría 🧵</h2>
            <h4 style="color: #0284C7; font-weight: 600; margin-bottom: 12px;">¡Con un clic transformamos tu logo en una obra de arte!</h4>
            <p style="font-size: 1rem; color: #4B5563; max-width: 700px; margin: 0 auto;">
                Bordados profesionales, uniformes escolares, corporativos y personalización textil de alta calidad con atención directa.
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )
    
    st.markdown("---")
    
    st.markdown("### 🎯 ¿Qué podemos hacer por ti?")
    
    col_serv1, col_serv2 = st.columns(2)
    
    with col_serv1:
        with st.container(border=True):
            st.markdown("#### 🏫 Uniformes y Colegios")
            try:
                st.image("colegio.jpg", use_container_width=True)
            except Exception:
                st.write("*(Foto de uniformes escolares)*")
                
            st.write("Ideal para el regreso a clases:")
            st.markdown("- 🛡️ Bordado de escudos institucionales\n- 👕 Playeras tipo polo escolares\n- 👔 Suéteres y camisas formales")
            
    with col_serv2:
        with st.container(border=True):
            st.markdown("#### 🏢 Empresas y Eventos")
            try:
                st.image("empresa.jpg", use_container_width=True)
            except Exception:
                st.write("*(Foto de ropa corporativa)*")
                
            st.write("Haz que tu marca destaque:")
            st.markdown("- 👔 Camisas empresariales y chalecos\n- 🧢 Gorras personalizadas\n- 🧵 Parches bordados y dotaciones")

    st.markdown("")
    st.info("💡 **¿Tienes tus propias prendas?** Tráelas o envíalas y nosotros nos encargamos de estampar o bordar tu logotipo con acabados de primera.")

    st.markdown("---")
    
    with st.container(border=True):
        st.markdown("### 🚀 ¿Listo para hacer un pedido o cotización?")
        st.write("Arma tu pedido paso a paso en segundos de forma rápida y sencilla.")
        
        if st.button("✨ Hacer un Pedido / Cotización Ahora", use_container_width=True, type="primary"):
            st.session_state.pagina_activa = "📝 Nueva Solicitud"
            st.session_state.paso = 1
            st.rerun()

# ==============================================================================
# MÓDULO 1: FORMULARIO CLIENTE (NUEVA SOLICITUD EN 4 PASOS)
# ==============================================================================
elif pagina == "📝 Nueva Solicitud":

    st.markdown("### 📝 Configura tu Solicitud")
    
    c_p1, c_p2, c_p3, c_p4 = st.columns(4)
    
    with c_p1:
        if st.button("1. Contacto", use_container_width=True, type="primary" if st.session_state.paso == 1 else "secondary"):
            st.session_state.paso = 1
            st.rerun()
    with c_p2:
        if st.button("2. Prendas", use_container_width=True, type="primary" if st.session_state.paso == 2 else "secondary"):
            if st.session_state.get("nombre"):
                st.session_state.paso = 2
                st.rerun()
            else:
                st.warning("Completa los datos de contacto primero.")
    with c_p3:
        if st.button("3. Personalización", use_container_width=True, type="primary" if st.session_state.paso == 3 else "secondary"):
            if st.session_state.get("colegios_agregados"):
                st.session_state.paso = 3
                st.rerun()
            else:
                st.warning("Agrega al menos un grupo de prendas primero.")
    with c_p4:
        if st.button("4. Resumen", use_container_width=True, type="primary" if st.session_state.paso == 4 else "secondary"):
            if st.session_state.get("colegios_agregados") and st.session_state.get("tipo_logo"):
                st.session_state.paso = 4
                st.rerun()
            else:
                st.warning("Completa los pasos anteriores.")

    st.divider()

    # PASO 1: DATOS DEL CLIENTE
    if st.session_state.paso == 1:
        st.progress(25)
        st.subheader("👤 1. Datos de Contacto")
        st.caption("Por favor, ingresa tus datos para gestionar tu solicitud de bordado.")

        with st.container(border=True):
            nombre = st.text_input("Nombre y Apellido *", value=st.session_state.get("nombre", ""), placeholder="Ej. Ana Mendoza")
            telefono = st.text_input("Teléfono de Contacto (WhatsApp) *", value=st.session_state.get("telefono", ""), placeholder="Ej. 04121234567")
            correo = st.text_input("Correo Electrónico *", value=st.session_state.get("correo", ""), placeholder="ejemplo@correo.com")

        if st.button("Continuar a Selección de Prendas ➡️", use_container_width=True, type="primary"):
            if not nombre.strip():
                st.error("Debe ingresar el nombre.")
            elif not telefono.strip() or len(telefono.strip()) < 10:
                st.error("Debe ingresar un teléfono válido (mínimo 10 dígitos).")
            elif "@" not in correo or "." not in correo:
                st.error("Debe ingresar un correo electrónico válido.")
            else:
                st.session_state.nombre = nombre
                st.session_state.telefono = telefono
                st.session_state.correo = correo
                st.session_state.paso = 2
                st.rerun()

    # PASO 2: SELECCIÓN DE COLEGIO Y PRENDAS
    elif st.session_state.paso == 2:
        st.progress(50)
        st.subheader("🏫 2. Colegio y Prendas")

        es_admin = (clave_admin == clave_correcta)

        df_col = obtener_colegios()
        lista_colegios = ["Seleccione un colegio (Opcional)..."] + (df_col["nombre"].dropna().tolist() if not df_col.empty else [])

        df_p = obtener_tipos_prenda()
        lista_tipos_prenda = ["Seleccione una prenda..."] + (df_p["nombre"].dropna().tolist() if not df_p.empty else [])

        df_t = obtener_tallas()
        lista_tallas = ["Seleccione una talla..."] + (df_t["nombre"].dropna().tolist() if not df_t.empty else [])

        df_m = obtener_marcas()
        lista_marcas = ["Seleccione una marca..."] + (df_m["nombre"].dropna().tolist() if not df_m.empty else [])

        df_c = obtener_colores()
        lista_colores = ["Seleccione un color..."] + (df_c["nombre"].dropna().tolist() if not df_c.empty else [])

        v = st.session_state.form_version

        with st.container(border=True):
            col_sel, col_btn = st.columns([4, 1] if es_admin else [1, 0.01])
            with col_sel:
                colegio = st.selectbox("Colegio o Institución", lista_colegios, key=f"colegio_sel_{v}")
            
            if es_admin:
                with col_btn:
                    st.write("")
                    with st.popover("➕"):
                        st.markdown("**Nuevo Colegio**")
                        nuevo_col_nom = st.text_input("Nombre", key=f"quick_col_nom_{v}")
                        nuevo_col_prec = st.number_input("Precio Bordado", min_value=0.0, step=0.5, key=f"quick_col_prec_{v}")
                        if st.button("Guardar", key=f"btn_quick_col_{v}"):
                            if nuevo_col_nom.strip():
                                guardar_colegio(nuevo_col_nom.strip(), nuevo_col_prec)
                                st.success("¡Colegio agregado!")
                                st.rerun()

            if colegio != "Seleccione un colegio (Opcional)..." and not df_col.empty:
                precio_base = obtener_precio_colegio(colegio)
                st.info(f"💡 Precio base de bordado para **{colegio}**: **${precio_base:.2f}** por prenda.")
            else:
                precio_base = 0.0

        st.subheader("👕 Detalle de la Prenda")

        with st.container(border=True):
            c_p1, c_p2 = st.columns([4, 1] if es_admin else [1, 0.01])
            with c_p1:
                tipo_prenda = st.selectbox("Tipo de Prenda *", lista_tipos_prenda, key=f"tipo_prenda_{v}")
            if es_admin:
                with c_p2:
                    st.write("")
                    with st.popover("➕"):
                        nuevo_p = st.text_input("Nueva Prenda", key=f"quick_p_{v}")
                        if st.button("Guardar", key=f"btn_quick_p_{v}") and nuevo_p.strip():
                            guardar_tipo_prenda(nuevo_p.strip())
                            st.rerun()

            col_a, col_b = st.columns(2)
            with col_a:
                c_t1, c_t2 = st.columns([3, 1] if es_admin else [1, 0.01])
                with c_t1:
                    talla = st.selectbox("Talla *", lista_tallas, key=f"talla_{v}")
                if es_admin:
                    with c_t2:
                        st.write("")
                        with st.popover("➕"):
                            nueva_t = st.text_input("Nueva Talla", key=f"quick_t_{v}")
                            if st.button("Guardar", key=f"btn_quick_t_{v}") and nueva_t.strip():
                                guardar_talla(nueva_t.strip())
                                st.rerun()

                c_c1, c_c2 = st.columns([3, 1] if es_admin else [1, 0.01])
                with c_c1:
                    color = st.selectbox("Color *", lista_colores, key=f"color_{v}")
                if es_admin:
                    with c_c2:
                        st.write("")
                        with st.popover("➕"):
                            nuevo_c = st.text_input("Nuevo Color", key=f"quick_c_{v}")
                            if st.button("Guardar", key=f"btn_quick_c_{v}") and nuevo_c.strip():
                                guardar_color(nuevo_c.strip())
                                st.rerun()

            with col_b:
                c_m1, c_m2 = st.columns([3, 1] if es_admin else [1, 0.01])
                with c_m1:
                    marca = st.selectbox("Marca *", lista_marcas, key=f"marca_{v}")
                if es_admin:
                    with c_m2:
                        st.write("")
                        with st.popover("➕"):
                            nueva_m = st.text_input("Nueva Marca", key=f"quick_m_{v}")
                            if st.button("Guardar", key=f"btn_quick_m_{v}") and nueva_m.strip():
                                guardar_marca(nueva_m.strip())
                                st.rerun()

                cantidad = st.number_input("Cantidad *", min_value=1, value=1, key=f"cantidad_{v}")

        if "prendas_actuales" not in st.session_state:
            st.session_state.prendas_actuales = []

        if st.button("➕ Agregar Prenda a la Lista Actual", use_container_width=True):
            if tipo_prenda == "Seleccione una prenda...":
                st.error("Debe seleccionar un tipo de prenda.")
            elif talla == "Seleccione una talla...":
                st.error("Debe seleccionar una talla.")
            elif marca == "Seleccione una marca...":
                st.error("Debe seleccionar una marca.")
            elif color == "Seleccione un color...":
                st.error("Debe seleccionar un color.")
            else:
                st.session_state.prendas_actuales.append({
                    "tipo": tipo_prenda,
                    "talla": talla,
                    "marca": marca,
                    "color": color,
                    "cantidad": cantidad
                })
                st.rerun()

        st.divider()
        st.subheader("📋 Prendas en preparación")

        if not st.session_state.prendas_actuales:
            st.info("Aún no has agregado prendas a este grupo.")
        else:
            for i, prenda in enumerate(st.session_state.prendas_actuales):
                c1, c2 = st.columns([5, 1])
                with c1:
                    st.success(f"👕 {prenda['tipo']} | 📏 {prenda['talla']} | 🏷️ {prenda['marca']} | 🎨 {prenda['color']} | 🔢 Cantidad: {prenda['cantidad']}")
                with c2:
                    if st.button("🗑️", key=f"borrar_prenda_{i}"):
                        st.session_state.prendas_actuales.pop(i)
                        st.rerun()

        if st.button("💾 Guardar Grupo de Prendas / Colegio", use_container_width=True, type="primary"):
            if not st.session_state.prendas_actuales:
                st.error("Debe agregar al menos una prenda antes de guardar.")
            else:
                nombre_grupo_colegio = colegio if colegio != "Seleccione un colegio (Opcional)..." else "General / Sin logo de Colegio"
                if any(c["colegio"] == nombre_grupo_colegio for c in st.session_state.colegios_agregados):
                    st.error("Ese grupo ya fue agregado a la lista.")
                else:
                    st.session_state.colegios_agregados.append({
                        "colegio": nombre_grupo_colegio,
                        "prendas": st.session_state.prendas_actuales.copy()
                    })
                    st.session_state.prendas_actuales = []
                    st.session_state.form_version += 1
                    st.success("✅ Grupo guardado en tu solicitud.")
                    st.rerun()

        st.divider()
        st.subheader("🏫 Grupos / Colegios Guardados en tu Solicitud")
        if not st.session_state.colegios_agregados:
            st.info("Aún no hay grupos listos en tu solicitud.")
        else:
            for idx_col, colegio_data in enumerate(st.session_state.colegios_agregados):
                with st.container(border=True):
                    c_lbl, c_btn = st.columns([5, 1])
                    with c_lbl:
                        st.markdown(f"#### 🛡️ {colegio_data['colegio']}")
                    with c_btn:
                        if st.button("🗑️ Eliminar Grupo", key=f"del_col_agregado_{idx_col}"):
                            st.session_state.colegios_agregados.pop(idx_col)
                            st.rerun()

                    for p_idx, prenda in enumerate(colegio_data["prendas"]):
                        st.write(f"&nbsp;&nbsp;&nbsp;&nbsp;• **{prenda['tipo']}** | Talla: {prenda['talla']} | Marca: {prenda['marca']} | Color: {prenda['color']} | Qty: **{prenda['cantidad']}**")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Volver a Datos", use_container_width=True):
                st.session_state.paso = 1
                st.rerun()
        with col2:
            if st.button("Continuar a Personalización ➡️", use_container_width=True, type="primary"):
                if not st.session_state.colegios_agregados:
                    st.error("Debe guardar al menos un grupo de prendas.")
                else:
                    st.session_state.paso = 3
                    st.rerun()

    # PASO 3: PERSONALIZACIÓN Y DELIVERY
    elif st.session_state.paso == 3:
        st.progress(75)
        st.subheader("🧵 3. Personalización y Entrega")

        total_prendas_pedido = sum(p["cantidad"] for c in st.session_state.colegios_agregados for p in c["prendas"])

        df_del = obtener_zonas_delivery()
        lista_zonas = []

        if not df_del.empty:
            col_zona = None
            for col in ["zona", "nombre", "zona_delivery"]:
                if col in df_del.columns:
                    col_zona = col
                    break
            
            if not col_zona and len(df_del.columns) > 0:
                col_zona = df_del.columns[0]

            if col_zona:
                lista_zonas = df_del[col_zona].dropna().tolist()

        with st.container(border=True):
            st.markdown("### 🎨 Opciones de Bordado")
            tipo_logo = st.selectbox(
                "Tipo de Logo / Arte *",
                ["Bordado Estándar del Colegio", "Personalizado / Diseñado por Cliente"],
                key="input_tipo_logo"
            )

            bordar_nombre = st.radio(
                "¿Desea bordar nombre personalizado en las prendas?",
                ["No", "Sí"],
                horizontal=True,
                key="input_bordar_nombre"
            )

            nombre_bordado = ""
            cantidad_nombre = 0
            if bordar_nombre == "Sí":
                nombre_bordado = st.text_input("Nombre / Texto a bordar (ej. Nombre y ubicación en la prenda)", placeholder="Ej. Juan Pérez", key="input_texto_nombre")
                cantidad_nombre = st.number_input(
                    f"¿En cuántas prendas se aplicará el nombre? (Máximo disponible: {total_prendas_pedido})", 
                    min_value=1, 
                    max_value=total_prendas_pedido, 
                    value=min(1, total_prendas_pedido), 
                    key="input_cant_nombre"
                )

        with st.container(border=True):
            st.markdown("### 🚚 Método de Entrega")
            delivery = st.radio(
                "¿Cómo desea recibir su pedido?",
                ["Retiro en Tienda", "Envío a Domicilio (Delivery)"],
                horizontal=True,
                key="input_delivery"
            )

            zona_delivery = "N/A"
            costo_delivery = 0.0

            if delivery == "Envío a Domicilio (Delivery)":
                if lista_zonas:
                    zona_delivery = st.selectbox("Seleccione su zona de delivery *", lista_zonas, key="input_zona")
                    costo_delivery = float(obtener_costo_delivery(zona_delivery) or 0.0)
                    st.info(f"🛵 Costo de envío a **{zona_delivery}**: **${costo_delivery:.2f}**")
                else:
                    st.warning("No hay zonas de delivery configuradas en el sistema.")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Volver a Prendas", use_container_width=True):
                st.session_state.paso = 2
                st.rerun()
        with col2:
            if st.button("Continuar al Resumen ➡️", use_container_width=True, type="primary"):
                if bordar_nombre == "Sí" and not nombre_bordado.strip():
                    st.error("Por favor, ingrese el nombre que desea bordar.")
                elif bordar_nombre == "Sí" and cantidad_nombre > total_prendas_pedido:
                    st.error(f"❌ La cantidad de nombres a bordar ({cantidad_nombre}) no puede ser mayor al total de prendas en tu pedido ({total_prendas_pedido}).")
                else:
                    st.session_state.tipo_logo = tipo_logo
                    st.session_state.bordar_nombre = bordar_nombre
                    st.session_state.nombre_bordado = nombre_bordado
                    st.session_state.cantidad_nombre = cantidad_nombre if bordar_nombre == "Sí" else 0
                    st.session_state.delivery = "Sí" if delivery == "Envío a Domicilio (Delivery)" else "No"
                    st.session_state.zona_delivery = zona_delivery
                    st.session_state.costo_delivery = costo_delivery
                    
                    st.session_state.paso = 4
                    st.rerun()

    # PASO 4: RESUMEN Y CONFIRMACIÓN DE PEDIDO
    elif st.session_state.paso == 4:
        st.progress(100)
        st.subheader("📋 4. Resumen Final de tu Solicitud")

        with st.container(border=True):
            st.markdown(f"### 👤 {st.session_state.nombre}")
            st.write(f"📞 **Teléfono:** {st.session_state.telefono}")
            st.write(f"📧 **Correo:** {st.session_state.correo}")

        with st.container(border=True):
            st.markdown("### 🧵 Detalles de Personalización")
            st.write(f"• **Tipo de Logo:** {st.session_state.tipo_logo}")
            st.write(f"• **Bordado de Nombre:** {st.session_state.bordar_nombre}")
            if st.session_state.bordar_nombre == "Sí":
                st.write(f"• **Detalle:** {st.session_state.nombre_bordado}")
                st.write(f"• **Prendas con nombre:** {st.session_state.cantidad_nombre}")

        dias_produccion = int(obtener_parametro("dias_produccion") or 3)
        fecha_entrega = date.today() + timedelta(days=dias_produccion)

        with st.container(border=True):
            st.markdown("### 💰 Presupuesto Estimado")
            st.info(f"📅 **Fecha estimada de entrega:** {fecha_entrega.strftime('%d/%m/%Y')}")

            subtotal_bordado = 0.0
            for colegio_data in st.session_state.colegios_agregados:
                colegio_nombre = colegio_data["colegio"]
                cantidad_colegio = sum(p["cantidad"] for p in colegio_data["prendas"])
                
                if colegio_nombre == "General / Sin logo de Colegio":
                    precio_colegio = 0.0
                else:
                    precio_colegio = obtener_precio_colegio(colegio_nombre)

                if cantidad_colegio >= 6 and precio_colegio > 0:
                    precio_colegio = max(0.0, precio_colegio - 0.50)

                subtotal_colegio = cantidad_colegio * precio_colegio
                subtotal_bordado += subtotal_colegio

                st.write(f"🏫 **{colegio_nombre}** {'🎉 *(Descuento de $0.50 aplicado por 6+ prendas)*' if cantidad_colegio >= 6 and precio_colegio > 0 else ''}")
                st.write(f"   ↳ {cantidad_colegio} prendas x ${precio_colegio:.2f} = **${subtotal_colegio:.2f}**")

            precio_nombre = float(obtener_parametro("precio_nombre") or 0)
            subtotal_nombres = (st.session_state.cantidad_nombre * precio_nombre) if st.session_state.bordar_nombre == "Sí" else 0.0

            if st.session_state.bordar_nombre == "Sí":
                st.write(f"🔤 **Nombres:** {st.session_state.cantidad_nombre} x ${precio_nombre:.2f} = **${subtotal_nombres:.2f}**")

            st.write(f"🚚 **Delivery ({st.session_state.zona_delivery if st.session_state.delivery == 'Sí' else 'Retiro en Tienda'}):** **${st.session_state.costo_delivery:.2f}**")
            
            total_estimado = subtotal_bordado + subtotal_nombres + st.session_state.costo_delivery
            st.divider()
            st.metric(label="Monto Total Estimado ($ USD)", value=f"${total_estimado:.2f}")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("⬅️ Modificar Datos", use_container_width=True):
                st.session_state.paso = 3
                st.rerun()
        with col2:
            if st.button("✅ Confirmar y Enviar Solicitud", key="confirmar_solicitud_mobile", use_container_width=True, type="primary", disabled=st.session_state.solicitud_enviada):
                colegio_orden = "Múltiples Grupos" if len(st.session_state.colegios_agregados) > 1 else st.session_state.colegios_agregados[0]["colegio"]
                cantidad_total = sum(p["cantidad"] for c in st.session_state.colegios_agregados for p in c["prendas"])

                orden_id = guardar_orden(
                    st.session_state.nombre,
                    st.session_state.telefono,
                    st.session_state.correo,
                    colegio_orden,
                    cantidad_total,
                    st.session_state.tipo_logo,
                    st.session_state.nombre_bordado,
                    st.session_state.cantidad_nombre,
                    st.session_state.delivery,
                    st.session_state.zona_delivery,
                    fecha_entrega,
                    0,
                    subtotal_bordado,
                    subtotal_nombres,
                    st.session_state.costo_delivery,
                    0,
                    total_estimado,
                    "Recibido"
                )

                for colegio_data in st.session_state.colegios_agregados:
                    for prenda in colegio_data["prendas"]:
                        guardar_detalle(
                            orden_id,
                            colegio_data["colegio"],
                            prenda["tipo"],
                            prenda["talla"],
                            prenda["marca"],
                            prenda["color"],
                            int(prenda["cantidad"])
                        )

                try:
                    enviar_confirmacion_solicitud(
                        st.session_state.correo,
                        st.session_state.nombre,
                        orden_id,
                        fecha_entrega
                    )
                except Exception as e:
                    st.error(f"Error enviando correo: {e}")

                st.session_state.solicitud_enviada = True
                st.session_state.ultimo_pedido = orden_id
                st.rerun()

        if st.session_state.solicitud_enviada:
            st.balloons()
            st.success(f"🎉 ¡Solicitud #{st.session_state.ultimo_pedido:04d} registrada con éxito!")
            st.info("Te hemos enviado un correo electrónico con la confirmación.")

            if st.button("➕ Crear otra Solicitud", key="nueva_solicitud_mobile", use_container_width=True):
                st.session_state.clear()
                st.rerun()

# ==============================================================================
# MÓDULO 2: CONSULTA DE ÓRDENES Y GESTIÓN ADMINISTRATIVA
# ==============================================================================
elif pagina == "📋 Consultas":
    st.title("📋 Consulta de Órdenes")
    df_ordenes = obtener_ordenes()

    if df_ordenes.empty:
        st.info("ℹ️ No hay órdenes registradas en la base de datos.")
    else:
        if HERRAMIENTAS_REPORTES_DISPONIBLES:
            try:
                excel_hist_file = generar_excel_historico(df_ordenes)
                with open(excel_hist_file, "rb") as f:
                    st.download_button(
                        "📊 Descargar Histórico General de Órdenes (Excel)",
                        f,
                        file_name=excel_hist_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="btn_descarga_historico_top"
                    )
            except Exception:
                pass

        st.divider()

        df_consulta = df_ordenes[["id", "nombre", "colegio", "status", "fecha_entrega", "saldo_pendiente", "fecha_pago"]].copy()
        df_consulta.columns = ["ID", "Cliente", "Colegio", "Estado", "Entrega", "Saldo", "Último Pago"]
        st.dataframe(df_consulta, use_container_width=True)

        pedido_id = st.selectbox("Seleccione un pedido", df_ordenes["id"].tolist())
        st.info(f"📦 Pedido seleccionado: #{pedido_id:04d}")

        pedido = obtener_orden_por_id(pedido_id).iloc[0]
        st.write(f"👤 {pedido['nombre']} | 📞 {pedido['telefono']} | 📧 {pedido['correo']}")

        detalle_orden = obtener_detalle_orden(pedido_id)
        with st.expander("👕 Prendas", expanded=True):
            if not detalle_orden.empty:
                st.dataframe(detalle_orden, use_container_width=True)
                
                col_cantidad = next((col for col in detalle_orden.columns if col.lower() == "cantidad"), None)
                
                if col_cantidad:
                    total_prendas_pedido = int(pd.to_numeric(detalle_orden[col_cantidad], errors="coerce").fillna(0).sum())
                    st.metric(
                        label="Total de prendas en este pedido",
                        value=f"{total_prendas_pedido} unidades"
                    )
                else:
                    st.warning("No se encontró la columna de cantidad en el detalle del pedido.")
            else:
                st.info("No hay prendas registradas para este pedido.")

        with st.expander("💰 Pagos y Tasa de Cambio (USD / Bs.)"):
            saldo = float(pedido.get("saldo_pendiente", 0))
            st.metric("Saldo Pendiente ($ USD)", f"${saldo:.2f}")

            tasa_actual = float(obtener_parametro("tasa_cambio") or 0.0)
            tasa_input = st.number_input("Tasa de Cambio (Bs / $)", min_value=0.0, value=tasa_actual, step=0.10, format="%.2f", key=f"tasa_pago_{pedido_id}")
            
            if tasa_input > 0 and saldo > 0:
                saldo_bs = saldo * tasa_input
                st.info(f"💡 **Saldo equivalente en Bolívares:** {saldo_bs:,.2f} Bs.")

            if pedido.get("fecha_pago"):
                st.caption(f"📅 Fecha del último pago registrado: {pedido.get('fecha_pago')}")

            if saldo <= 0:
                st.success("💳 Estado Pago: ✅ Pagado")
            else:
                st.warning(f"💳 Estado Pago: 🔴 Pendiente (${saldo:.2f})")

            monto_pago = st.number_input("Monto recibido ($ USD)", min_value=0.0, step=1.0, key=f"monto_in_{pedido_id}")
            
            if monto_pago > 0 and tasa_input > 0:
                monto_bs = round(monto_pago * tasa_input, 2)
                st.write(f"💵 **Equivalente del pago:** {monto_bs:,.2f} Bs.")

            if st.button("💾 Registrar Pago", key=f"reg_pago_{pedido_id}"):
                if monto_pago > saldo and saldo > 0:
                    st.error("❌ El pago excede el saldo pendiente.")
                elif monto_pago <= 0:
                    st.error("❌ Monto inválido.")
                else:
                    registrar_pago(pedido_id, monto_pago, tasa_input)
                    st.success("✅ Pago y registro en Bolívares guardados correctamente.")
                    st.rerun()

            df_hist_pagos = obtener_historico_pagos(pedido_id)
            if not df_hist_pagos.empty:
                st.subheader("📜 Histórico de Pagos en Bolívares")
                st.dataframe(df_hist_pagos[["monto_usd", "tasa_cambio", "monto_bs", "fecha"]], use_container_width=True)

        with st.expander("📄 Documentos"):
            tasa_doc = st.number_input("Tasa para Comprobante PDF (Opcional)", min_value=0.0, value=float(obtener_parametro("tasa_cambio") or 0.0), step=0.10, key=f"tasa_pdf_{pedido_id}")
            
            if st.button("📄 Generar PDF"):
                if HERRAMIENTAS_REPORTES_DISPONIBLES:
                    pdf_file = generar_pdf_orden(pedido, detalle_orden, tasa_doc)
                    with open(pdf_file, "rb") as f:
                        st.download_button("📥 Descargar PDF", f, file_name=pdf_file, mime="application/pdf")
                else:
                    st.error("El módulo de generación de PDFs no está disponible.")

            if st.button("📊 Generar Excel Pedido"):
                if HERRAMIENTAS_REPORTES_DISPONIBLES:
                    excel_file = generar_excel_orden(pedido, detalle_orden)
                    with open(excel_file, "rb") as f:
                        st.download_button("📥 Descargar Excel Pedido", f, file_name=excel_file, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.error("El módulo de generación de Excel no está disponible.")

        with st.expander("📧 Comunicaciones"):
            if st.button("📧 Reenviar Correo"):
                pdf_file = "comprobante.pdf"
                if HERRAMIENTAS_REPORTES_DISPONIBLES:
                    pdf_file = generar_pdf_orden(pedido, detalle_orden)
                enviar_pdf_por_correo(pedido["correo"], pedido["nombre"], pedido["id"], pedido["fecha_entrega"], pdf_file)
                st.success("✅ Correo reenviado")

        with st.expander("🏭 Producción"):
            nuevo_estado = st.selectbox("Cambiar estado", ["Recibido", "En Producción", "Listo para Entrega", "Anulado"])
            if st.button("🔄 Actualizar Estado"):
                actualizar_status_orden(pedido_id, nuevo_estado)
                enviar_notificacion_estado(pedido["correo"], pedido["nombre"], pedido["id"], pedido["fecha_entrega"], nuevo_estado, pedido["delivery"])
                st.success("✅ Estado actualizado correctamente")
                st.rerun()

        with st.expander("🗑️ Zona de Peligro / Eliminar Pedido"):
            st.error("⚠️ **Atención:** Esta acción es permanente. Se eliminará la orden, sus ítems y sus pagos registrados.")
            
            confirmar = st.checkbox(f"Estoy seguro de que deseo eliminar la Orden #{pedido_id:04d}", key=f"chk_confirm_{pedido_id}")
            
            if st.button("🗑️ Eliminar Pedido Definitivamente", key=f"btn_del_{pedido_id}"):
                if confirmar:
                    if eliminar_orden(pedido_id):
                        st.success(f"✅ La Orden #{pedido_id:04d} ha sido eliminada correctamente.")
                        st.rerun()
                    else:
                        st.error("❌ Ocurrió un error al intentar borrar el pedido de la base de datos.")
                else:
                    st.warning("⚠️ Marca la casilla de verificación anterior para confirmar la eliminación.")

# ==============================================================================
# MÓDULO ADMINISTRATIVO: DASHBOARD, NÓMINA, REPORTES FASES B Y C Y EXCEL
# ==============================================================================
elif pagina == "📊 Reportes":
    import io
    import pandas as pd
    import plotly.express as px
    import plotly.graph_objects as go
    from datetime import date, timedelta
    import streamlit as st
    
    # Librerías para dar estilo profesional a Excel
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # --------------------------------------------------------------------------
    # FUNCIONES DE FASE B: GENERACIÓN DE EXCEL HISTÓRICO CON FORMATO CONDICIONAL Y TOTALES
    # --------------------------------------------------------------------------
    def generar_excel_historico(df_ordenes):
        buffer = io.BytesIO()
        if df_ordenes.empty:
            return buffer

        df_export = df_ordenes.copy()

        cols_deseadas = [c for c in ["id", "nombre", "telefono", "email", "colegio", "cantidad_total", "delivery", "delivery_costo", "status", "monto_total", "abono", "saldo_pendiente", "fecha_entrega"] if c in df_export.columns]
        df_export = df_export[cols_deseadas]

        if "id" in df_export.columns:
            df_export["id"] = df_export["id"].apply(lambda x: f"#{int(x):04d}" if str(x).isdigit() else str(x))

        renombres = {
            "id": "ID Pedido", "nombre": "Cliente", "telefono": "Teléfono", "email": "Correo",
            "colegio": "Colegio / Detalle", "cantidad_total": "Cant. Prendas", "delivery": "Delivery",
            "delivery_costo": "Costo Delivery ($)", "status": "Estado", "monto_total": "Monto Total ($)",
            "abono": "Abonado ($)", "saldo_pendiente": "Saldo Pendiente ($)", "fecha_entrega": "Fecha Entrega"
        }
        df_export.rename(columns=renombres, inplace=True)

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name="Historico_General", index=False, startrow=2)
            wb = writer.book
            ws = wb["Historico_General"]

            # Estilos Base
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Segoe UI", size=13, bold=True, color="1F4E79")
            total_font = Font(name="Segoe UI", size=11, bold=True)
            body_font = Font(name="Segoe UI", size=10)
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )
            double_bottom_border = Border(
                top=Side(style='thin', color='000000'),
                bottom=Side(style='double', color='000000')
            )

            # Banner de Título Superior
            ws.merge_cells("A1:M1")
            cell_title = ws["A1"]
            cell_title.value = "HISTÓRICO GENERAL DE ÓRDENES Y PAGOS - BORDACLICK"
            cell_title.font = title_font
            cell_title.fill = title_fill
            cell_title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35

            # Encabezados
            ws.row_dimensions[3].height = 25
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=3, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Formato Condicional por Estado
            fill_completado = PatternFill(start_color="E2EFDA", fill_type="solid")  # Verde
            fill_proceso = PatternFill(start_color="FFF2CC", fill_type="solid")     # Amarillo
            fill_pendiente = PatternFill(start_color="FCE4D6", fill_type="solid")   # Naranja

            num_filas = len(df_export)
            fila_inicio = 4
            fila_fin = fila_inicio + num_filas - 1

            for row in range(fila_inicio, fila_fin + 1):
                ws.row_dimensions[row].height = 20
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = body_font
                    cell.border = thin_border
                    
                    col_name = str(ws.cell(row=3, column=col).value)

                    if col_name == "Estado" and cell.value:
                        val_lower = str(cell.value).lower()
                        if any(k in val_lower for k in ["entregado", "listo", "pagado", "completado"]):
                            cell.fill = fill_completado
                        elif any(k in val_lower for k in ["producción", "borde", "proceso"]):
                            cell.fill = fill_proceso
                        elif any(k in val_lower for k in ["recibido", "pendiente"]):
                            cell.fill = fill_pendiente

                    if any(term in col_name for term in ["($)", "Monto", "Abonado", "Saldo"]):
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif any(term in col_name for term in ["ID", "Cant", "Delivery", "Fecha"]):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Fila de Totales Automáticos (Fórmulas Excel SUM)
            fila_total = fila_fin + 1
            ws.row_dimensions[fila_total].height = 24
            
            ws.cell(row=fila_total, column=1, value="TOTALES GENERALES").font = total_font
            ws.cell(row=fila_total, column=1).alignment = Alignment(horizontal="left", vertical="center")

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=fila_total, column=col)
                cell.fill = total_fill
                cell.border = double_bottom_border
                col_letter = get_column_letter(col)
                col_name = str(ws.cell(row=3, column=col).value)

                if "Cant. Prendas" in col_name:
                    cell.value = f"=SUM({col_letter}{fila_inicio}:{col_letter}{fila_fin})"
                    cell.font = total_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif any(term in col_name for term in ["($)", "Monto", "Abonado", "Saldo"]):
                    cell.value = f"=SUM({col_letter}{fila_inicio}:{col_letter}{fila_fin})"
                    cell.font = total_font
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Auto-ajuste de Ancho de Columnas
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1: continue
                    if cell.value is not None:
                        val_str = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer.seek(0)
        return buffer

    # --------------------------------------------------------------------------
    # INICIO DE LA INTERFAZ STREAMLIT
    # --------------------------------------------------------------------------
    st.title("📊 Dashboard y Reportes de Administración")
    st.caption("🔒 Módulo exclusivo de control administrativo")
    
    df_ordenes_rep = obtener_ordenes()
    
    if df_ordenes_rep.empty:
        st.info("ℹ️ No hay datos registrados en la base de datos para generar reportes.")
    else:
        # 0. MAPEO Y REPARACIÓN DE DATOS
        mapeo_columnas = {
            'monto': 'monto_total', 'total': 'monto_total', 'precio_total': 'monto_total',
            'monto_orden': 'monto_total', 'precio': 'monto_total', 'delivery_precio': 'delivery_costo',
            'costo_delivery': 'delivery_costo', 'saldo': 'saldo_pendiente', 'abono_inicial': 'abono',
            'cantidad': 'cantidad_total', 'estado': 'status'
        }
        df_ordenes_rep.rename(columns=mapeo_columnas, inplace=True)

        columnas_base = {
            'monto_total': 0.0, 'abono': 0.0, 'saldo_pendiente': 0.0, 'delivery_costo': 0.0,
            'cantidad_total': 0, 'status': 'Pendiente', 'colegio': 'N/A', 'delivery': 'No',
            'email': '', 'correo': '', 'telefono': '', 'fecha_pago': None, 'fecha_entrega': None, 'fecha_creacion': None
        }
        
        for col, val_defecto in columnas_base.items():
            if col not in df_ordenes_rep.columns:
                df_ordenes_rep[col] = val_defecto

        for col in ['monto_total', 'abono', 'saldo_pendiente', 'delivery_costo', 'cantidad_total']:
            df_ordenes_rep[col] = pd.to_numeric(df_ordenes_rep[col], errors='coerce').fillna(0)

        # Reparación si el monto_total viene en 0
        mask_cero = (df_ordenes_rep['monto_total'] <= 0)
        df_ordenes_rep.loc[mask_cero, 'monto_total'] = df_ordenes_rep.loc[mask_cero, 'abono'] + df_ordenes_rep.loc[mask_cero, 'saldo_pendiente']

        # Priorizar FECHA DE PAGO sobre fecha de creación/entrega
        df_ordenes_rep['fecha_liquidada'] = df_ordenes_rep['fecha_pago'].fillna(df_ordenes_rep['fecha_entrega']).fillna(df_ordenes_rep['fecha_creacion'])
        df_ordenes_rep['fecha_dt'] = pd.to_datetime(df_ordenes_rep['fecha_liquidada'], errors='coerce').dt.date

        # 1. FILTRO DE FECHAS
        with st.container(border=True):
            st.markdown("### 📅 Seleccionar Período (Nómina por Fecha de Cobro / Pago)")
            hoy = date.today()
            hace_7_dias = hoy - timedelta(days=6)
            
            col_f_fecha, col_f_quick = st.columns([2, 1])
            with col_f_fecha:
                rango_fechas = st.date_input(
                    "Rango de fechas de cobro (Desde - Hasta):",
                    value=(hace_7_dias, hoy),
                    key="filtro_rango_semanal"
                )
            with col_f_quick:
                st.write("")
                st.caption("💡 Incluye todas las órdenes pagadas/liquidadas en este rango, sin importar su fecha de origen.")

        if isinstance(rango_fechas, (list, tuple)) and len(rango_fechas) == 2:
            f_inicio, f_fin = rango_fechas[0], rango_fechas[1]
            df_periodo = df_ordenes_rep[
                (df_ordenes_rep['fecha_dt'] >= f_inicio) & 
                (df_ordenes_rep['fecha_dt'] <= f_fin)
            ].copy()
            st.success(f"🗓️ Evaluando órdenes pagadas/liquidadas desde **{f_inicio.strftime('%d/%m/%Y')}** hasta **{f_fin.strftime('%d/%m/%Y')}**")
        else:
            f_inicio, f_fin = hace_7_dias, hoy
            df_periodo = df_ordenes_rep.copy()

        # 2. SECCIÓN DE NÓMINA (BORDADOR Y DELIVERY)
        df_pagadas = df_periodo[
            (df_periodo["saldo_pendiente"] <= 0) | 
            (df_periodo["status"].astype(str).str.lower().isin(["pagado", "entregado", "completado"]))
        ].copy()

        st.markdown("---")
        st.subheader("💵 Resumen de Pagos a Delivery y Bordador (Órdenes Cobradas)")

        df_delivery_pagados = df_pagadas[
            df_pagadas["delivery"].astype(str).str.lower().isin(["sí", "si", "true", "1", "delivery", "con delivery"])
        ]

        total_viajes_delivery = len(df_delivery_pagados)
        total_monto_delivery = df_delivery_pagados["delivery_costo"].sum()
        subtotal_neto_bordado = (df_pagadas["monto_total"] - df_pagadas["delivery_costo"]).clip(lower=0).sum()

        with st.expander("🧮 Calculadora de Pago al Bordador (Porcentaje sobre Venta Neta)", expanded=True):
            pct_bordador = st.number_input(
                "Porcentaje (%) a pagar al bordador sobre el total neto cobrado (sin delivery):", 
                min_value=0.0, max_value=100.0, value=25.0, step=0.5, key="pct_empleado_bordado"
            )
            total_nomina_bordador = subtotal_neto_bordado * (pct_bordador / 100.0)

        col_pay1, col_pay2, col_pay3, col_pay4 = st.columns(4)
        with col_pay1: st.metric("🛵 Viajes Delivery", f"{total_viajes_delivery} viajes")
        with col_pay2: st.metric("📦 Total Repartidor", f"${total_monto_delivery:,.2f}")
        with col_pay3: st.metric("🧵 Base Neto Bordado", f"${subtotal_neto_bordado:,.2f}")
        with col_pay4: st.metric("💰 TOTAL BORDADOR", f"${total_nomina_bordador:,.2f}", help=f"{pct_bordador}% sobre la base neta")

        # EXCEL DE NÓMINA Y DELIVERIES
        buffer_nomina = io.BytesIO()
        with pd.ExcelWriter(buffer_nomina, engine='openpyxl') as writer:
            cols_nom = [c for c in ["id", "nombre", "telefono", "email", "colegio", "cantidad_total", "delivery", "delivery_costo", "status", "fecha_liquidada", "monto_total", "abono", "saldo_pendiente"] if c in df_pagadas.columns]
            df_nom_export = df_pagadas[cols_nom].copy()
            
            if "id" in df_nom_export.columns:
                df_nom_export["id"] = df_nom_export["id"].apply(lambda x: f"#{int(x):04d}" if str(x).isdigit() else str(x))

            df_nom_export["Base Neto Bordado ($)"] = (df_nom_export["monto_total"] - df_nom_export["delivery_costo"]).clip(lower=0)
            df_nom_export[f"Pago Bordador ({pct_bordador}%) ($)"] = df_nom_export["Base Neto Bordado ($)"] * (pct_bordador / 100.0)

            nombres_cols = {
                "id": "ID Pedido", "nombre": "Cliente", "telefono": "Teléfono", "email": "Correo",
                "colegio": "Colegio", "cantidad_total": "Cant. Prendas", "delivery": "Delivery",
                "delivery_costo": "Costo Delivery ($)", "status": "Estado", "fecha_liquidada": "Fecha de Cobro/Pago",
                "monto_total": "Monto Total ($)", "abono": "Abonado ($)", "saldo_pendiente": "Saldo Pendiente ($)"
            }
            df_nom_export.rename(columns=nombres_cols, inplace=True)
            df_nom_export.to_excel(writer, sheet_name="Detalle_Nomina", index=False, startrow=2)

            df_resumen_nomina = pd.DataFrame([
                {"CONCEPTO": "Período de Liquidación Desde", "VALOR / MONTO": str(f_inicio)},
                {"CONCEPTO": "Período de Liquidación Hasta", "VALOR / MONTO": str(f_fin)},
                {"CONCEPTO": "Total Órdenes Liquidadas", "VALOR / MONTO": len(df_pagadas)},
                {"CONCEPTO": "Monto Total Facturado", "VALOR / MONTO": df_pagadas["monto_total"].sum()},
                {"CONCEPTO": "Total Descontado por Delivery", "VALOR / MONTO": total_monto_delivery},
                {"CONCEPTO": "Base Neta para Bordado", "VALOR / MONTO": subtotal_neto_bordado},
                {"CONCEPTO": "Porcentaje Comisión Aplicado", "VALOR / MONTO": f"{pct_bordador}%"},
                {"CONCEPTO": "TOTAL A PAGAR AL BORDADOR", "VALOR / MONTO": total_nomina_bordador},
                {"CONCEPTO": "TOTAL A PAGAR AL REPARTIDOR", "VALOR / MONTO": total_monto_delivery}
            ])
            df_resumen_nomina.to_excel(writer, sheet_name="Resumen_Pago", index=False, startrow=2)

            wb = writer.book
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            title_font = Font(name="Segoe UI", size=14, bold=True, color="1F4E79")
            body_font = Font(name="Segoe UI", size=10)
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )

            for ws in wb.worksheets:
                titulo_texto = "HISTÓRICO DE ÓRDENES LIQUIDADAS Y COBRADAS - BORDACLICK" if ws.title == "Detalle_Nomina" else "RESUMEN EJECUTIVO DE NÓMINA Y PAYROLL"
                ws.merge_cells("A1:M1" if ws.title == "Detalle_Nomina" else "A1:B1")
                cell_title = ws["A1"]
                cell_title.value = titulo_texto
                cell_title.font = title_font
                cell_title.fill = title_fill
                cell_title.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 35

                ws.row_dimensions[3].height = 26
                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=3, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for row in range(4, ws.max_row + 1):
                    ws.row_dimensions[row].height = 20
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = body_font
                        cell.border = thin_border
                        
                        col_name = str(ws.cell(row=3, column=col).value)
                        
                        if any(term in col_name for term in ["($)", "Monto", "Abonado", "Saldo", "Base", "Pago", "TOTAL"]):
                            cell.number_format = '"$"#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif "ID" in col_name or "Cant" in col_name or "Delivery" in col_name or "Fecha" in col_name:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row == 1: continue
                        if cell.value is not None:
                            val_str = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value)
                            max_len = max(max_len, len(val_str))
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buffer_nomina.seek(0)

        st.download_button(
            label="📄 Descargar Recibo de Nómina y Delivery Profesional (.xlsx)",
            data=buffer_nomina,
            file_name=f"nomina_bordaclick_{f_inicio}_al_{f_fin}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.markdown("---")

        # 3. INDICADORES CLAVE (KPIs GENERALES)
        st.subheader("📌 Indicadores Clave del Período")
        tot_pedidos = len(df_periodo)
        tot_abonos = df_periodo['abono'].sum()
        tot_saldos = df_periodo['saldo_pendiente'].sum()
        tot_facturado = tot_abonos + tot_saldos
        ticket_prom = tot_facturado / tot_pedidos if tot_pedidos > 0 else 0

        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        with kpi1: st.metric("Total Facturado ($)", f"${tot_facturado:,.2f}")
        with kpi2: st.metric("💵 Total Recaudado ($)", f"${tot_abonos:,.2f}")
        with kpi3: st.metric("⏳ Cuentas por Cobrar ($)", f"${tot_saldos:,.2f}")
        with kpi4: st.metric("🏷️ Ticket Promedio ($)", f"${ticket_prom:,.2f}")

        st.markdown("---")

        # 4. GRÁFICOS ESTADÍSTICOS (PLOTLY)
        st.subheader("📈 Análisis de Operaciones del Período")
        g_col1, g_col2 = st.columns(2)

        with g_col1:
            df_st = df_periodo["status"].value_counts().reset_index()
            df_st.columns = ["Estado", "Cantidad"]
            fig_status = px.pie(
                df_st, values="Cantidad", names="Estado", 
                title="Distribución de Órdenes por Estado", hole=0.4,
                color_discrete_sequence=px.colors.qualitative.Set2
            )
            fig_status.update_traces(textposition="inside", textinfo="percent+label")
            st.plotly_chart(fig_status, use_container_width=True)

        with g_col2:
            fig_balance = go.Figure(data=[
                go.Bar(name="Cobrado ($)", x=["Balance"], y=[tot_abonos], marker_color="#2E7D32"),
                go.Bar(name="Por Cobrar ($)", x=["Balance"], y=[tot_saldos], marker_color="#C62828")
            ])
            fig_balance.update_layout(barmode="group", title="Balance Financiero ($ USD)", yaxis_title="Monto ($)", height=380)
            st.plotly_chart(fig_balance, use_container_width=True)

        st.markdown("---")

        # ----------------------------------------------------------------------
        # FASE C: REPORTES ESPECIALIZADOS DE OPERACIÓN (TALLER Y LOGÍSTICA)
        # ----------------------------------------------------------------------
        st.subheader("🛠️ Reportes Especializados de Operación")
        
        tab_taller, tab_logistica = st.tabs(["🧵 Reporte de Taller (Producción)", "🛵 Reporte de Logística (Delivery)"])

        # 1. REPORTE DE TALLER / PRODUCCIÓN
        with tab_taller:
            st.markdown("#### 🪡 Hoja de Trabajo para Bordadores y Ensamblaje")
            st.caption("Filtro simplificado listo para imprimir o enviar al taller (sin datos de precios/cobros).")

            df_taller = df_periodo[
                df_periodo["status"].astype(str).str.lower().isin(["recibido", "en producción", "pendiente", "en proceso"])
            ].copy()

            if df_taller.empty:
                st.info("ℹ️ No hay órdenes pendientes de bordado o en producción actualmente.")
            else:
                tot_prendas_taller = df_taller["cantidad_total"].sum()
                
                c1, c2 = st.columns(2)
                with c1: st.metric("📋 Órdenes Activas en Taller", f"{len(df_taller)} pedidos")
                with c2: st.metric("🧵 Total Prendas a Bordar", f"{tot_prendas_taller} piezas")

                cols_taller = [c for c in ["id", "nombre", "colegio", "cantidad_total", "status", "fecha_entrega"] if c in df_taller.columns]
                df_taller_view = df_taller[cols_taller].rename(columns={
                    "id": "ID Pedido", "nombre": "Cliente", "colegio": "Diseño / Colegio",
                    "cantidad_total": "Cant. Prendas", "status": "Estado Actual", "fecha_entrega": "Fecha Compromiso"
                })

                st.dataframe(df_taller_view, use_container_width=True)

                buf_taller = io.BytesIO()
                with pd.ExcelWriter(buf_taller, engine='openpyxl') as writer:
                    df_taller_view.to_excel(writer, sheet_name="Taller_Produccion", index=False)
                buf_taller.seek(0)

                st.download_button(
                    label="🖨️ Descargar Hoja de Taller (.xlsx)",
                    data=buf_taller,
                    file_name="reporte_taller_produccion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        # 2. REPORTE DE LOGÍSTICA / DELIVERY
        with tab_logistica:
            st.markdown("#### 📦 Guía de Despacho y Rutas de Entrega")
            st.caption("Ficha de ruta optimizada para el motorizado / servicio de reparto.")

            df_logistica = df_periodo[
                df_periodo["delivery"].astype(str).str.lower().isin(["sí", "si", "true", "1", "delivery", "con delivery"])
            ].copy()

            if df_logistica.empty:
                st.info("ℹ️ No hay entregas por delivery programadas en el período seleccionado.")
            else:
                col_log1, col_log2 = st.columns(2)
                with col_log1: st.metric("🛵 Total Envíos a Domicilio", f"{len(df_logistica)} entregas")
                with col_log2: st.metric("💵 Total Cobro Delivery", f"${df_logistica['delivery_costo'].sum():,.2f}")

                cols_log = [c for c in ["id", "nombre", "telefono", "colegio", "status", "delivery_costo", "saldo_pendiente", "fecha_entrega"] if c in df_logistica.columns]
                df_log_view = df_logistica[cols_log].rename(columns={
                    "id": "ID Pedido", "nombre": "Cliente", "telefono": "Teléfono Contacto",
                    "colegio": "Detalle / Referencia", "status": "Estado Delivery",
                    "delivery_costo": "Flete ($)", "saldo_pendiente": "Cobrar en Destino ($)", "fecha_entrega": "Fecha Entrega"
                })

                st.dataframe(df_log_view, use_container_width=True)

                buf_log = io.BytesIO()
                with pd.ExcelWriter(buf_log, engine='openpyxl') as writer:
                    df_log_view.to_excel(writer, sheet_name="Ruta_Delivery", index=False)
                buf_log.seek(0)

                st.download_button(
                    label="🛵 Descargar Guía para Delivery (.xlsx)",
                    data=buf_log,
                    file_name="guia_rutas_delivery.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        st.markdown("---")

        # ----------------------------------------------------------------------
        # 5. FILTROS EN PANTALLA Y DESCARGA HISTÓRICA COMPLETA (FASE B)
        # ----------------------------------------------------------------------
        with st.container(border=True):
            st.subheader("🔍 Filtros Adicionales de la Tabla General")
            col_f1, col_f2 = st.columns(2)
            with col_f1:
                estados_disponibles = ["Todos"] + list(df_periodo["status"].dropna().unique())
                filtro_estado = st.selectbox("Filtrar por Estado", estados_disponibles)
            with col_f2:
                colegios_disponibles = ["Todos"] + list(df_periodo["colegio"].dropna().unique())
                filtro_colegio = st.selectbox("Filtrar por Colegio", colegios_disponibles)

        df_filtrado = df_periodo.copy()
        if filtro_estado != "Todos": df_filtrado = df_filtrado[df_filtrado["status"] == filtro_estado]
        if filtro_colegio != "Todos": df_filtrado = df_filtrado[df_filtrado["colegio"] == filtro_colegio]

        st.dataframe(df_filtrado, use_container_width=True)

        # Generar el Excel Histórico FASE B (con Totales automáticos y Formato Condicional)
        buffer_historico = generar_excel_historico(df_filtrado)

        st.markdown("### 📥 Descargar Reporte Histórico General")
        st.download_button(
            label="📊 Descargar Histórico General con Totales y Formato Condicional (.xlsx)",
            data=buffer_historico,
            file_name=f"historico_general_bordaclick_{f_inicio}_al_{f_fin}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
# ------------------------------------------------------------------------------
# SECCIONES ADMINISTRATIVAS: CONFIGURACIÓN Y CATÁLOGOS
# ------------------------------------------------------------------------------

elif pagina == "⚙️ Configuración":
    st.title("⚙️ Configuración General")
    
    try:
        precio_actual = obtener_parametro("precio_bordado_nombre") if 'obtener_parametro' in globals() else 2.0
        dias_actual = int(obtener_parametro("dias_produccion") if 'obtener_parametro' in globals() else 3)
    except Exception as e:
        st.error(f"Error al conectar con la base de datos para cargar parámetros: {e}")
        precio_actual = 2.0
        dias_actual = 3

    with st.container(border=True):
        precio_bordado_nombre = st.number_input(
            "Precio Bordado de Nombre",
            min_value=0.0,
            step=0.5,
            value=float(precio_actual if precio_actual is not None else 2.0),
            key="cfg_precio_nombre"
        )
        
        dias_produccion = st.number_input(
            "Días de Producción",
            min_value=1,
            step=1,
            value=int(dias_actual if dias_actual is not None else 3),
            key="cfg_dias_produccion"
        )
        
        if st.button("💾 Guardar Configuración", use_container_width=True, key="btn_save_config"):
            try:
                guardar_parametro("precio_bordado_nombre", str(precio_bordado_nombre))
                guardar_parametro("dias_produccion", str(dias_produccion))
                st.success("✅ Configuración guardada con éxito.")
                st.rerun()
            except Exception as e:
                st.error(f"No se pudo guardar la configuración: {e}")

elif pagina == "🏫 Colegios":
    st.title("🏫 Gestión de Colegios")
    
    with st.container(border=True):
        nombre_colegio = st.text_input("Nombre del Colegio", key="admin_nom_colegio")
        precio_colegio = st.number_input("Precio Bordado Colegio ($ USD)", min_value=0.0, step=0.50, key="admin_precio_colegio")
        
        if st.button("💾 Guardar Colegio", use_container_width=True, key="btn_save_colegio"):
            if nombre_colegio.strip():
                guardar_colegio(nombre_colegio.strip(), precio_colegio)
                st.success(f"✅ Colegio '{nombre_colegio}' registrado con éxito.")
                st.rerun()
            else:
                st.warning("⚠️ El nombre del colegio no puede estar vacío.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    
    df_colegios = obtener_colegios() if 'obtener_colegios' in globals() else pd.DataFrame()
    
    if not df_colegios.empty:
        st.dataframe(df_colegios, use_container_width=True)
        
        with st.expander("🗑️ Eliminar un Colegio"):
            opciones_col = {f"{row['nombre']} ($ {row['precio_bordado']})": row['id'] for _, row in df_colegios.iterrows()}
            seleccion_col = st.selectbox("Seleccione el colegio a eliminar:", list(opciones_col.keys()), key="del_col_select")
            
            if st.button("🗑️ Eliminar Colegio", key="btn_del_col"):
                eliminar_colegio(opciones_col[seleccion_col])
                st.success("✅ Colegio eliminado con éxito.")
                st.rerun()
    else:
        st.info("No hay colegios registrados o la tabla está vacía.")

elif pagina == "🚚 Delivery":
    st.title("🚚 Gestión de Zonas de Delivery")
    
    with st.container(border=True):
        zona = st.text_input("Nombre de la Zona / Sector", key="admin_zona_deliv")
        precio_delivery = st.number_input("Costo de Envío ($ USD)", min_value=0.0, step=0.50, key="admin_precio_deliv")
        if st.button("💾 Guardar Zona", use_container_width=True, key="btn_save_deliv"):
            if zona.strip():
                guardar_zona_delivery(zona, precio_delivery) if 'guardar_zona_delivery' in globals() else None
                st.success(f"✅ Zona '{zona}' guardada.")
                st.rerun()
            else:
                st.warning("⚠️ Ingrese el nombre de la zona.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    df_delivery = obtener_zonas_delivery() if 'obtener_zonas_delivery' in globals() else pd.DataFrame()
    if not df_delivery.empty:
        st.dataframe(df_delivery, use_container_width=True)
        with st.expander("🗑️ Eliminar una Zona"):
            opciones = {f"{row.get('nombre', 'Zona')} (ID: {row.get('id', 0)})": row.get('id', 0) for _, row in df_delivery.iterrows()}
            seleccion = st.selectbox("Seleccione la zona a eliminar:", list(opciones.keys()), key="del_deliv_select")
            if st.button("🗑️ Eliminar Zona", key="btn_del_deliv"):
                if 'eliminar_zona_delivery' in globals():
                    eliminar_zona_delivery(opciones[seleccion])
                    st.success("✅ Zona eliminada con éxito.")
                    st.rerun()
    else:
        st.info("No hay zonas de delivery registradas.")

elif pagina == "📦 Prendas":
    st.title("📦 Gestión de Tipos de Prenda")
    
    with st.container(border=True):
        tipo_prenda = st.text_input("Nombre del Tipo de Prenda", key="admin_tipo_prenda")
        if st.button("💾 Guardar Tipo de Prenda", use_container_width=True, key="btn_save_prenda"):
            if tipo_prenda.strip():
                guardar_tipo_prenda(tipo_prenda) if 'guardar_tipo_prenda' in globals() else None
                st.success(f"✅ Prenda '{tipo_prenda}' guardada.")
                st.rerun()
            else:
                st.warning("⚠️ Ingrese el nombre de la prenda.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    df_prendas = obtener_tipos_prenda() if 'obtener_tipos_prenda' in globals() else pd.DataFrame()
    if not df_prendas.empty:
        st.dataframe(df_prendas, use_container_width=True)
        with st.expander("🗑️ Eliminar un Tipo de Prenda"):
            opciones = {f"{row.get('nombre', 'Prenda')} (ID: {row.get('id', 0)})": row.get('id', 0) for _, row in df_prendas.iterrows()}
            seleccion = st.selectbox("Seleccione la prenda a eliminar:", list(opciones.keys()), key="del_prenda_select")
            if st.button("🗑️ Eliminar Prenda", key="btn_del_prenda"):
                if 'eliminar_tipo_prenda' in globals():
                    eliminar_tipo_prenda(opciones[seleccion])
                    st.success("✅ Prenda eliminada con éxito.")
                    st.rerun()
    else:
        st.info("No hay tipos de prenda registrados.")

elif pagina == "🏷️ Marcas":
    st.title("🏷️ Gestión de Marcas")
    
    with st.container(border=True):
        marca = st.text_input("Nombre de la Marca", key="admin_marca")
        if st.button("💾 Guardar Marca", use_container_width=True, key="btn_save_marca"):
            if marca.strip():
                guardar_marca(marca) if 'guardar_marca' in globals() else None
                st.success(f"✅ Marca '{marca}' guardada.")
                st.rerun()
            else:
                st.warning("⚠️ Ingrese el nombre de la marca.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    df_marcas = obtener_marcas() if 'obtener_marcas' in globals() else pd.DataFrame()
    if not df_marcas.empty:
        st.dataframe(df_marcas, use_container_width=True)
        with st.expander("🗑️ Eliminar una Marca"):
            opciones = {f"{row.get('nombre', 'Marca')} (ID: {row.get('id', 0)})": row.get('id', 0) for _, row in df_marcas.iterrows()}
            seleccion = st.selectbox("Seleccione la marca a eliminar:", list(opciones.keys()), key="del_marca_select")
            if st.button("🗑️ Eliminar Marca", key="btn_del_marca"):
                if 'eliminar_marca' in globals():
                    eliminar_marca(opciones[seleccion])
                    st.success("✅ Marca eliminada con éxito.")
                    st.rerun()
    else:
        st.info("No hay marcas registradas.")

elif pagina == "📏 Tallas":
    st.title("📏 Gestión de Tallas")
    
    with st.container(border=True):
        talla = st.text_input("Identificador de Talla", key="admin_talla")
        if st.button("💾 Guardar Talla", use_container_width=True, key="btn_save_talla"):
            if talla.strip():
                guardar_talla(talla) if 'guardar_talla' in globals() else None
                st.success(f"✅ Talla '{talla}' guardada.")
                st.rerun()
            else:
                st.warning("⚠️ Ingrese una talla.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    df_tallas = obtener_tallas() if 'obtener_tallas' in globals() else pd.DataFrame()
    if not df_tallas.empty:
        st.dataframe(df_tallas, use_container_width=True)
        with st.expander("🗑️ Eliminar una Talla"):
            opciones = {f"{row.get('nombre', 'Talla')} (ID: {row.get('id', 0)})": row.get('id', 0) for _, row in df_tallas.iterrows()}
            seleccion = st.selectbox("Seleccione la talla a eliminar:", list(opciones.keys()), key="del_talla_select")
            if st.button("🗑️ Eliminar Talla", key="btn_del_talla"):
                if 'eliminar_talla' in globals():
                    eliminar_talla(opciones[seleccion])
                    st.success("✅ Talla eliminada con éxito.")
                    st.rerun()
    else:
        st.info("No hay tallas registradas.")

elif pagina == "🎨 Colores":
    st.title("🎨 Gestión de Colores")
    
    with st.container(border=True):
        color = st.text_input("Nombre del Color", key="admin_color")
        if st.button("💾 Guardar Color", use_container_width=True, key="btn_save_color"):
            if color.strip():
                guardar_color(color) if 'guardar_color' in globals() else None
                st.success(f"✅ Color '{color}' guardado.")
                st.rerun()
            else:
                st.warning("⚠️ Ingrese un color.")

    st.divider()
    st.subheader("📋 Registros Actuales")
    df_colores = obtener_colores() if 'obtener_colores' in globals() else pd.DataFrame()
    if not df_colores.empty:
        st.dataframe(df_colores, use_container_width=True)
        with st.expander("🗑️ Eliminar un Color"):
            opciones = {f"{row.get('nombre', 'Color')} (ID: {row.get('id', 0)})": row.get('id', 0) for _, row in df_colores.iterrows()}
            seleccion = st.selectbox("Seleccione el color a eliminar:", list(opciones.keys()), key="del_color_select")
            if st.button("🗑️ Eliminar Color", key="btn_del_color"):
                if 'eliminar_color' in globals():
                    eliminar_color(opciones[seleccion])
                    st.success("✅ Color eliminado con éxito.")
                    st.rerun()
    else:
        st.info("No hay colores registrados.")

elif pagina == "💾 Respaldo":
    st.title("💾 Respaldo de Base de Datos")
    with st.container(border=True):
        try:
            with open("bordaclick_dev.db", "rb") as db_file:
                st.download_button(
                    label="📥 Descargar Base de Datos (bordaclick_dev.db)",
                    data=db_file,
                    file_name="bordaclick_dev_backup.db",
                    mime="application/x-sqlite3",
                    use_container_width=True,
                    key="btn_download_db"
                )
        except Exception as e:
            st.error(f"❌ Error al leer la base de datos: {e}")